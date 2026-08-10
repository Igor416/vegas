import re
from pathlib import Path

from django.conf import settings
from django.db import transaction
from openpyxl import load_workbook
from rest_framework.response import Response
from rest_framework.views import APIView

from api.models import Product, Size

EXCEL_PATH = Path(settings.BASE_DIR) / "temp.xlsx"
GAPS = ((9, 22), (24, 37))
PRODUCT_FIRST_COL = 3
PRODUCT_LAST_COL = 23
DIMENSION_RE = re.compile(r"^(\d+)\s*[xх×]\s*(\d+)$", re.IGNORECASE)

workbook = load_workbook(EXCEL_PATH, data_only=True)
sheet = workbook.active


def parse_dimension(value):
    if value is None:
        return None
    match = DIMENSION_RE.match(str(value).strip())
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def price_mdl(excel_price):
    return round(float(excel_price))


class WorkerView(APIView):
    def get(self, request):
        products_by_name = {
            product.name_en: product
            for product in Product.objects.filter(category__name="mattress")
        }
        sizes_by_key = {
            (size.product_id, size.width, size.length): size
            for size in Size.objects.filter(product__category__name="mattress")
        }

        updates = {}

        for gap_start, gap_end in GAPS:
            column_products = {}
            for col in range(PRODUCT_FIRST_COL, PRODUCT_LAST_COL + 1):
                name = sheet.cell(gap_start, col).value
                if not name:
                    continue
                product = products_by_name.get(str(name).strip())
                if product:
                    column_products[col] = product

            for row in range(gap_start + 1, gap_end + 1):
                dimension = parse_dimension(sheet.cell(row, 1).value)
                if not dimension:
                    continue

                width, length = dimension
                for col, product in column_products.items():
                    raw_price = sheet.cell(row, col).value
                    size = sizes_by_key.get((product.id, width, length))
                    if not size:
                        continue

                    if raw_price is None:
                        size.disabled = True
                    else:
                        size.priceMDL = price_mdl(raw_price)
                        size.disabled = False

                    updates[size.id] = size

        with transaction.atomic():
            Size.objects.bulk_update(updates.values(), ["priceMDL", "disabled"])

        return Response({"updated": len(updates)})
