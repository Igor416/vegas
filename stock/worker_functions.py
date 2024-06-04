from stock.models import Table, Stockable, Action
from openpyxl import load_workbook
from django.utils.timezone import datetime

def load_all():
  wb = load_workbook(r'C:\Users\User\Downloads\Telegram Desktop\оба два (Автосохраненный) (1).xlsx', data_only=True)
  ws = wb.active
  t = Table.objects.get(month=5)
  for row in range(3, 206):
    if row in [41, 75, 119, 159, 196]:
      continue
    product = ws.cell(row, 1).value
    size = ws.cell(row, 2).value
    allowed = lambda x: x != None and (type(x) == int or x.isnumeric())
    maping = {'4': 'S1', '5': 'S2', '7': 'C', '8': 'J', '9': 'R'}
    qs = Stockable.objects.filter(product=product, size=size)
    '''
    i = 0
    for col in range(14, 44):
      if allowed(ws.cell(row, col).value):
        for _ in range(ws.cell(row, col).value):
          while qs[i].current_place != 'S1' and qs[i].last_action.type != 'T':
            i+=1
          qs[i].actions.add(Action.objects.create(type='S', person='admin', place='S1', date=datetime(2024, m(ws.cell(2, col).value), ws.cell(2, col).value).strftime("%Y-%m-%d"), stockable=qs[i]))
          i+=1
    '''
    '''
    i = 0
    for col in range(4, 10):
      if col == 6 or not allowed(ws.cell(row, col).value):
        continue
      for _ in range(int(ws.cell(row, col).value)):
        s = Stockable.objects.create(product=product, size=size, table=t)
        s.actions.add(Action.objects.create(type='A', person='admin', place=maping[str(col)], date=datetime(2024, 5, 1).strftime("%Y-%m-%d"), stockable=s))
        i += 1
    '''
    '''
    i = 0
    if type(ws.cell(row, 12).value) == str and '+' in ws.cell(row, 12).value:
      for _ in range(int(ws.cell(row, 12).value.split('+')[1])):
        while qs[i].current_place != 'S1':
          i+=1
        qs[i].actions.add(Action.objects.create(type='O', person='admin', place='S1', date=datetime(2024, 5, 1).strftime("%Y-%m-%d"), stockable=qs[i]))
        i += 1
    '''
    '''
    items = sum(map(int, filter(allowed, [ws.cell(row, col).value for col in range(4, 44) if col != 12])))
    for i in range(items):
      s = Stockable.objects.create(product=product, size=size)
      s.actions.add(
        Action.objects.create(type='A', person='admin', place='Машина', date=datetime(2024, 4, 1).strftime("%Y-%m-%d"), stockable=s),
        Action.objects.create(type='T', person='admin', place='S1', date=datetime(2024, 4, 1).strftime("%Y-%m-%d"), stockable=s),
      )
    '''

def load_init():
  wb = load_workbook(r'C:\Users\User\Downloads\Telegram Desktop\06.xlsx', data_only=True)
  ws = wb.active
  table = Table.objects.get(month=6)
  for row in range(77, 90):
    product = str(ws.cell(row, 3).value).strip()
    size = ws.cell(row, 2).value
    if size.count('x') == 2:
      size = 'x'.join(size.split('x')[:-1])
    print(product, size)
    #'''
    for i in range(int(ws.cell(row, 4).value)):
      s = Stockable.objects.create(product=product, size=size, table=table)
      s.actions.add(Action.objects.create(type='A', person='admin', place='Машина', date=datetime(2024, 6, 1).strftime("%Y-%m-%d"), stockable=s))
    #'''