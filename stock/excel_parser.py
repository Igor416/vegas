from django.utils.timezone import datetime
from stock.models import Table, Stockable, Action
from openpyxl import load_workbook

def parse():
  hidden_table= Table.objects.get(year=2034, month=6)
  wb = load_workbook(r'C:\Users\User\Downloads\Telegram Desktop\оба два (Автосохраненный).xlsx', data_only=True)
  ws = wb.active
  for row in range(3, 325):
    name = ws.cell(row, 1).value
    size = ws.cell(row, 2).value
    if not name:
      name = ''
    name = str(name).strip()
    if size and 'x' in size:
      width, length = list(map(int, size.split('x')))
      found = Stockable.objects.filter(product=name, width=width, length=length)
      if found.exists():
        continue
        #found = found.first()
        places = {
          4: 'S1',
          7: 'C',
          8: 'J',
          9: 'R'
        }
        in_stock = []
        for place in range(4, 10):
          if place not in places.keys():
            continue
          count = ws.cell(row, place).value
          if not count:
            count = 0
          count = int(count)
          for _ in range(count):
            #s = Stockable.objects.create(category=found.category, product=name, width=width, length=length, table=hidden_table)
            s = Stockable.objects.create(product=name, width=width, length=length, table=hidden_table)
            s.actions.add(Action.objects.create(type='A', person='admin', place=places[place], stockable=s))
            if place == 4:
              in_stock.append(s)
        total = str(ws.cell(row, 11).value)
        if '+' in total:
          ordered = int(total.split('+')[1])
          for i in range(ordered):
            in_stock[i].actions.add(Action.objects.create(type='O', person='admin', place=places[4], stockable=in_stock[i]))
        for date in range(13, 25):
          count = ws.cell(row, date).value
          if not count:
            count = 0
          count = int(count)
          for _ in range(count):
            #s = Stockable.objects.create(category=found.category, product=name, width=width, length=length, table=hidden_table)
            s = Stockable.objects.create(product=name, width=width, length=length, table=hidden_table)
            s.actions.add(Action.objects.create(type='S', person='admin', place='S1', date=datetime(2024, 6, date - 3), stockable=s))
      else:
        print('unfound', name, size)
    else:
      print('error', name, size)